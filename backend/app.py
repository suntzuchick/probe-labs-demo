import io
import os
import re
import sys
import traceback

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "extractors"))

from dotenv import load_dotenv
# Loads backend/.env if present (local dev convenience — gitignored, never
# committed). In production (Render) env vars are set directly in the
# dashboard and this is a no-op since no .env file exists there.
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

from flask import Flask, request, jsonify, g
from flask_cors import CORS
import pandas as pd

from agent import extract
from extraction_trace import build_trace
import session_store as store
import derive_adam
import derive_contextual
import notebook_engine
import validate as _validate
import auth as _auth
import indexer
import dashboard_engine
import synthesis_engine
import understanding_engine
import db
import dataset_registry
import evidence
import norm_registry
import narrative_workspace
import oracle_engine
import reports
import narrative_engine
import corpus_store
import hypothesis_registry
import editor_lint

from flask import redirect, send_from_directory

_FRONTEND_ROOT = os.path.join(os.path.dirname(__file__), "..", "frontend")
# The single-page pipeline wizard (Ingest -> Extraction -> Quality -> Derive ->
# Notebook -> Oracles -> Analysis -> Evidence -> Narratives) is the product's
# real UI. It's plain HTML/CSS/JS with no build step, so Flask serves it
# directly rather than a Vite dist/ output.
FRONTEND_DIR = os.path.join(_FRONTEND_ROOT, "legacy")

app = Flask(__name__)
CORS(app, expose_headers=["X-Probe-Token"])
db.init_db()


@app.errorhandler(Exception)
def handle_unexpected_error(e):
    # Without this, an unhandled exception in a Claude-calling route (real LLM
    # output doesn't always match the shape a mocked test assumed) returns
    # Flask's raw HTML 500 page. The frontend's res.json() then throws on a
    # parse error, which surfaces as nothing happening at all rather than a
    # readable error. Print the full traceback server-side (terminal running
    # `python3 app.py`) and return real JSON so the frontend can show it.
    traceback.print_exc()
    return jsonify({"status": "error", "error": f"{type(e).__name__}: {e}"}), 500


@app.after_request
def _persist_touched_sessions(response):
    # Write-through every session this request read or mutated, regardless
    # of which route it was. Session state otherwise lives only in this
    # worker's memory (session_store._sessions) and is lost on any restart;
    # relying on individual routes to remember db.persist_session() left
    # long stretches of the pipeline (derive, quality, dashboards, analysis)
    # unpersisted, so "unknown session" showed up as soon as the process
    # recycled mid-session.
    touched = getattr(g, "_touched_sessions", None)
    if touched:
        for sid in touched:
            sess = store.peek(sid)
            if sess is not None:
                db.persist_session(sid, sess)
    return response


def _session_token_from_request():
    return request.headers.get("X-Probe-Token") or request.args.get("probe_token")


@app.before_request
def enforce_auth():
    if not _auth.auth_enabled():
        return
    if not request.path.startswith("/api/"):
        return
    if request.path.startswith("/api/auth/"):
        return
    token = _session_token_from_request()
    if not _auth.validate_session(token or ""):
        return jsonify({"error": "unauthorized", "code": "auth_required"}), 401


@app.route("/")
def serve_index():
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.route("/<path:filename>")
def serve_static(filename):
    return send_from_directory(FRONTEND_DIR, filename)


@app.route("/api/auth/request", methods=["POST"])
def auth_request_link():
    body  = request.get_json(force=True)
    email = (body.get("email") or "").strip().lower()
    if not email or "@" not in email:
        return jsonify({"error": "A valid email address is required."}), 400

    if not _auth.auth_enabled():
        return jsonify({"status": "sent"})

    if not _auth.is_allowed(email):
        return jsonify({"status": "not_allowed"})

    token    = _auth.create_magic_token(email)
    base_url = os.environ.get("APP_BASE_URL", "").rstrip("/")
    link     = f"{base_url}/api/auth/verify?token={token}"
    return jsonify({"status": "ok", "link": link})


@app.route("/api/auth/verify", methods=["GET"])
def auth_verify():
    token = request.args.get("token", "")
    session_token = _auth.verify_magic_token(token)
    if not session_token:
        return (
            "<!DOCTYPE html><html><body style='font-family:Arial,sans-serif;padding:48px;color:#333;max-width:480px;margin:auto'>"
            "<h2 style='color:#4E0C70'>Link expired or already used</h2>"
            "<p>Magic links expire after 7 days and can only be used once.</p>"
            "<p><a href='/' style='color:#2289FA'>← Request a new link</a></p>"
            "</body></html>"
        ), 400

    base = os.environ.get("APP_BASE_URL", "").rstrip("/")
    return redirect(f"{base}/?probe_token={session_token}", code=302)


@app.route("/api/auth/status", methods=["GET"])
def auth_status():
    token = _session_token_from_request()
    email = _auth.validate_session(token or "") if token else None
    return jsonify({
        "authenticated": bool(email),
        "email": email or "",
        "auth_enabled": _auth.auth_enabled(),
    })


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    token = _session_token_from_request()
    if token:
        _auth.revoke_session(token)
    return jsonify({"status": "ok"})

SAMPLE_DATA_DIR = os.path.join(os.path.dirname(__file__), "sample_data")
NOTEBOOK_STATE_ROOT = os.path.join(os.path.dirname(__file__), "notebook_state")


def _session_dir(sid):
    return os.path.join(NOTEBOOK_STATE_ROOT, sid)


def _slug(name: str) -> str:
    import re as _re
    return _re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_") or "data"


def _auto_save_domain(sid: str, filename: str, content: bytes, extraction_result: dict):
    fmt = extraction_result.get("detected_format", "")
    sdir = _session_dir(sid)
    os.makedirs(sdir, exist_ok=True)
    sess = store.get_session(sid)

    if fmt == "xlsx" and extraction_result.get("sheets"):
        _save_xlsx_sheets(filename, content, extraction_result, sdir, sess)
        return

    domain = derive_contextual.auto_detect_domain(extraction_result)
    if not domain:
        return

    try:
        sep = "\t" if fmt == "tsv" else ","
        df = pd.read_csv(io.BytesIO(content), sep=sep, dtype=str, keep_default_na=False)
    except Exception:
        return

    if df is None or df.empty:
        return

    var_name = _slug(filename.rsplit(".", 1)[0]) if domain == "DATA" else domain.lower()
    notebook_engine.save_state(sdir, var_name, df)

    mappings = {
        m["source_label"]: {"top_match": m["top_match"], "confidence": m["confidence"], "action": m["action"]}
        for m in extraction_result.get("mapping", [])
    }
    sess["domain_data"][domain] = []
    sess["domain_source"][domain] = {
        "filename": filename, "var_name": var_name,
        "n_rows": len(df), "auto_detected": True, "mappings": mappings,
    }


def _save_xlsx_sheets(filename, content, extraction_result, sdir, sess):
    from sdtm_mapping import is_plate_layout
    from tabular import _reshape_plate

    try:
        xls = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
    except Exception:
        return

    sheet_meta = {s["sheet_name"]: s for s in extraction_result.get("sheets", []) if s.get("status") == "ok"}

    for sheet_name in xls.sheet_names:
        meta = sheet_meta.get(sheet_name)
        if not meta:
            continue
        try:
            raw_df = xls.parse(sheet_name, dtype=str, keep_default_na=False)
            if raw_df.empty:
                continue

            cols = [str(c) for c in raw_df.columns]
            raw_df.columns = cols

            if meta.get("detected_layout") == "plate_map" and is_plate_layout(cols, raw_df[cols[0]].tolist()[:20]):
                df = _reshape_plate(raw_df)
            else:
                df = raw_df

            var_name = _slug(sheet_name)
            notebook_engine.save_state(sdir, var_name, df)

            mappings = {
                m["source_label"]: {"top_match": m["top_match"], "confidence": m["confidence"], "action": m["action"]}
                for m in meta.get("mapping", [])
            }
            domain_key = "SHEET_" + _slug(sheet_name).upper()
            sess["domain_data"][domain_key] = []
            sess["domain_source"][domain_key] = {
                "filename": filename, "sheet_name": sheet_name, "var_name": var_name,
                "n_rows": len(df), "layout": meta.get("detected_layout", "long_format"),
                "auto_detected": True, "mappings": mappings,
            }
        except Exception:
            continue


@app.route("/api/session", methods=["POST"])
def create_session():
    sid = store.new_session()
    os.makedirs(_session_dir(sid), exist_ok=True)
    return jsonify({"session_id": sid})


@app.route("/api/session/info", methods=["POST"])
def session_info():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    # Merge, don't replace — a caller updating just "project" (e.g. renaming
    # a workspace) shouldn't blank out name/organization/email set earlier.
    existing = dict(sess.get("user_info") or {})
    for field in ("name", "organization", "email", "project"):
        if field in body:
            existing[field] = body[field].strip() if isinstance(body[field], str) else body[field]
    sess["user_info"] = existing
    return jsonify({"status": "ok"})


@app.route("/api/session/list", methods=["GET"])
def session_list():
    """Real workspace switcher backing — every persisted session, summarized.
    A "workspace" in this UI is a project-scoped session (its own datasets,
    narratives, evidence), not the narrative-drafting workspace/branching
    unit narrative_workspace.py implements — those are two different things
    that happen to share a name in the design this maps from."""
    out = []
    for s in db.list_sessions():
        sid = s["sid"]
        out.append({
            "sid": sid,
            "name": (s.get("user_info") or {}).get("project") or f"Session {sid[:8]}",
            "narratives": len(db.list_narratives(sid, include_synthetic=False)),
            "datasets": len((dataset_registry.get_current(sid) or {}).get("manifest", {})),
            "evidence": len(db.list_evidence(sid)),
            "updated_at": s.get("created_at"),
            "current": sid == request.args.get("current_sid"),
        })
    return jsonify({"status": "ok", "workspaces": out})


@app.route("/api/session/<sid>/status", methods=["GET"])
def session_status(sid):
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    domain_rows = {d: len(sess["domain_data"][d]) for d in sess["domain_data"]}
    plan = derive_contextual.get_plan(sess)
    return jsonify({
        "session_id": sid,
        "files_uploaded": list(sess["files"].keys()),
        "domains_mapped": list(sess["domain_data"].keys()),
        "domain_row_counts": domain_rows,
        "context": plan["context"],
        "context_label": plan["context_label"],
        "notebook_vars": notebook_engine.available_vars(_session_dir(sid)),
        "oracles_enabled": bool(sess.get("oracles_enabled", False)),
        "act": sess.get("act", 1),
    })


@app.route("/api/upload", methods=["POST"])
def upload():
    sid = request.form.get("session_id")
    if not sid or not store.session_exists(sid):
        return jsonify({"error": "missing or unknown session_id"}), 400
    if "file" not in request.files:
        return jsonify({"error": "no file in request"}), 400

    f = request.files["file"]
    filename = f.filename
    content = f.read()

    result = extract(filename, content)
    trace = build_trace(result)

    sess = store.get_session(sid)
    sess["files"][filename] = result

    _auto_save_domain(sid, filename, content, result)

    return jsonify({"extraction": result, "trace": trace})


@app.route("/api/load-sample", methods=["POST"])
def load_sample():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404

    sess = store.get_session(sid)
    domain_files = {"DM": "DM.csv", "EX": "EX.csv", "AE": "AE.csv", "RS": "RS.csv", "DS": "DS.csv", "DV": "DV.csv"}
    loaded = {}
    all_traces = []

    for domain, fname in domain_files.items():
        path = os.path.join(SAMPLE_DATA_DIR, fname)
        with open(path, "rb") as fh:
            content = fh.read()
        extraction = extract(fname, content)
        trace = build_trace(extraction)
        all_traces.append({"filename": fname, "domain": domain, "trace": trace})

        sess["files"][fname] = extraction
        df = pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False)
        full_rows = df.to_dict(orient="records")
        sess["domain_data"][domain] = full_rows
        sess["domain_source"][domain] = {"filename": fname, "n_rows": len(full_rows)}

        notebook_engine.save_state(_session_dir(sid), domain.lower(), df)

        loaded[domain] = {"filename": fname, "n_rows": len(full_rows)}

    return jsonify({
        "status": "ok", "loaded": loaded, "traces": all_traces,
    })


@app.route("/api/confirm-mapping", methods=["POST"])
def confirm_mapping():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    filename = body.get("filename")
    domain = body.get("domain")
    column_map = body.get("column_map", {})
    sheet_name = body.get("sheet_name")

    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404

    sess = store.get_session(sid)
    if filename not in sess["files"]:
        return jsonify({"error": f"file '{filename}' was not uploaded in this session"}), 400

    file_result = sess["files"][filename]

    if file_result.get("detected_format") == "xlsx":
        sheet = next((s for s in file_result["sheets"] if s["sheet_name"] == sheet_name), None)
        if sheet is None:
            return jsonify({"error": f"sheet '{sheet_name}' not found"}), 400
        preview_rows = sheet.get("preview", [])
    else:
        preview_rows = file_result.get("preview", [])

    if not preview_rows:
        return jsonify({"error": "no row data available to map"}), 400

    renamed_rows = [{column_map.get(k, k): v for k, v in row.items()} for row in preview_rows]

    sess["domain_data"][domain] = renamed_rows
    sess["domain_source"][domain] = {"filename": filename, "sheet_name": sheet_name, "n_rows": len(renamed_rows)}

    df = pd.DataFrame(renamed_rows)
    notebook_engine.save_state(_session_dir(sid), domain.lower(), df)

    return jsonify({
        "status": "ok", "domain": domain, "n_rows_mapped": len(renamed_rows),
        "missing_domains": store.missing_domains(sid),
    })


def _load_all_dfs(sdir: str) -> dict:
    dfs = {}
    for v in notebook_engine.available_vars(sdir):
        df = notebook_engine.load_state(sdir, v)
        # Code Builder cells can leave behind non-DataFrame scratch values
        # (lists, scalars, tuples) under names that were once DataFrames —
        # the kernel persists on name reuse regardless of type. Only real
        # tables belong in the indexed dataset view.
        if isinstance(df, pd.DataFrame):
            dfs[v] = df
    return dfs


TRIAL_SUFFIXES = ("a", "b", "c")


def _load_triple_dfs(sdir: str) -> dict:
    """Only the triple-trial suffixed tables (dm_a, adsl_a, adsl_b, adsl_c,
    ...) — Act 2/3 sessions may still carry an earlier Act 1's unprefixed
    tables in the same sandbox dir (one continuous session), and those
    shouldn't leak into cross-trial schema context."""
    dfs = {}
    for v in notebook_engine.available_vars(sdir):
        if not any(v.endswith(f"_{s}") for s in TRIAL_SUFFIXES):
            continue
        df = notebook_engine.load_state(sdir, v)
        if isinstance(df, pd.DataFrame):
            dfs[v] = df
    return dfs


TRIPLE_SAMPLE_DIRS = {
    "a": (SAMPLE_DATA_DIR, "RASOLUTE302-SIM", "Trial A — RASolute 302"),
    "b": (os.path.join(SAMPLE_DATA_DIR, "trial_b"), "RASOLUTE-CONFIRM-B", "Trial B — confirmatory, larger, multi-region"),
    "c": (os.path.join(SAMPLE_DATA_DIR, "trial_c"), "RASOLUTE-EXPAND-C", "Trial C — expansion cohort, smaller N"),
}


@app.route("/api/load-triple-trials", methods=["POST"])
def load_triple_trials():
    """Act 2/3's ingestion step — three prebaked clinical trials (same drug
    class, different populations/effect sizes) loaded in one call so the
    demo can go straight to real cross-trial comparison. Real extraction/
    trace machinery is reused per trial, just namespaced by suffix."""
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    sdir = _session_dir(sid)

    domain_files = {"DM": "DM.csv", "EX": "EX.csv", "AE": "AE.csv", "RS": "RS.csv", "DS": "DS.csv", "DV": "DV.csv"}
    trials = {}
    all_traces = []
    for suffix, (data_dir, study_id, label) in TRIPLE_SAMPLE_DIRS.items():
        loaded = {}
        for domain, fname in domain_files.items():
            path = os.path.join(data_dir, fname)
            with open(path, "rb") as fh:
                content = fh.read()
            extraction = extract(fname, content)
            trace = build_trace(extraction)
            all_traces.append({"filename": f"{fname} ({label})", "domain": domain, "trace": trace, "trial": suffix})

            df = pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False)
            notebook_engine.save_state(sdir, f"{domain.lower()}_{suffix}", df)
            loaded[domain] = {"filename": fname, "n_rows": len(df)}
        trials[suffix] = {"label": label, "study_id": study_id, "loaded": loaded}

    sess["act"] = 2
    sess["trials"] = trials
    return jsonify({"status": "ok", "trials": trials, "traces": all_traces})


@app.route("/api/derive-triple", methods=["POST"])
def derive_triple():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    if not sess.get("trials"):
        return jsonify({"status": "error", "error": "Load the three trials first."}), 200

    results = {}
    for suffix in TRIAL_SUFFIXES:
        try:
            dm = notebook_engine.load_state(sdir, f"dm_{suffix}")
            ex = notebook_engine.load_state(sdir, f"ex_{suffix}")
            ae = notebook_engine.load_state(sdir, f"ae_{suffix}")
            rs = notebook_engine.load_state(sdir, f"rs_{suffix}")
            ds = notebook_engine.load_state(sdir, f"ds_{suffix}")
            dv = notebook_engine.load_state(sdir, f"dv_{suffix}")
            adsl, adae, adtte = derive_adam.run_pipeline(dm, ex, ae, rs, ds, dv=dv)
        except Exception as e:
            results[suffix] = {"status": "error", "error": str(e)}
            continue
        derived = {"adsl": adsl, "adae": adae, "adtte": adtte, "rs": rs}
        if dv is not None and not dv.empty:
            derived["addv"] = derive_adam.derive_addv(dv, adsl)
        for key, df in derived.items():
            notebook_engine.save_state(sdir, f"{key}_{suffix}", df)
        results[suffix] = {
            "status": "ok",
            "datasets": {f"{key}_{suffix}": {"rows": len(df), "columns": list(df.columns)} for key, df in derived.items()},
        }

    sess.setdefault("derived_vars", [])
    for suffix, r in results.items():
        if r["status"] == "ok":
            sess["derived_vars"] = sorted(set(sess["derived_vars"]) | set(r["datasets"].keys()))

    return jsonify({"status": "ok", "trials": results})


@app.route("/api/index/build-triple", methods=["POST"])
def index_build_triple():
    """Combined Contextual Understanding across all 3 trials — runs the same
    deterministic classifier per trial (indexer.classify_dataset already
    expects unprefixed table names, so each trial's suffixed tables get
    temporarily de-suffixed for the call) and merges the results."""
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    triple_dfs = _load_triple_dfs(sdir)
    if not triple_dfs:
        return jsonify({"status": "error", "error": "No triple-trial data derived yet."}), 200

    per_trial = {}
    entities, metrics, supported, unsupported, risks = set(), set(), set(), set(), []
    for suffix in TRIAL_SUFFIXES:
        sub_dfs = {name[:-len(f"_{suffix}")]: df for name, df in triple_dfs.items() if name.endswith(f"_{suffix}")}
        if not sub_dfs:
            continue
        u = indexer.classify_dataset(sub_dfs, "clinical_trial")
        label = (sess.get("trials", {}).get(suffix) or {}).get("label", f"Trial {suffix.upper()}")
        u["label"] = label
        u["n_subjects"] = int(sub_dfs["adsl"]["USUBJID"].nunique()) if "adsl" in sub_dfs and "USUBJID" in sub_dfs["adsl"].columns else None
        per_trial[suffix] = u
        entities |= set(u["entities"])
        metrics |= set(u["available_metrics"])
        supported |= set(u["supported_analyses"])
        unsupported |= set(u["unsupported_analyses"])
        risks.extend(f"{label}: {r}" for r in u.get("risks", []))

    combined = {
        "dataset_type": "clinical_trial_adam_multi",
        "trials": per_trial,
        "entities": sorted(entities), "available_metrics": sorted(metrics),
        "supported_analyses": sorted(supported), "unsupported_analyses": sorted(unsupported - supported),
        "risks": risks, "generated_by": "heuristic",
    }
    schema = indexer._schema_index(triple_dfs)

    sess["dataset_understanding"] = combined
    if not sess.get("indexes"):
        sess["indexes"] = {}
    sess["indexes"]["schema"] = schema
    sess["indexes"]["dataset"] = {"supported": combined["supported_analyses"], "unsupported": combined["unsupported_analyses"]}

    return jsonify({"status": "ok", "understanding": combined})


@app.route("/api/understanding/candidates", methods=["GET"])
def understanding_candidates():
    """Contextual Understanding's own gen-AI pass — the Data Understanding
    Agent asks questions about the data's nature/provenance/comparability
    (never a statistic — that's Notebook's job), works the same for a single
    dataset or the combined multi-trial view, whichever this session has."""
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    understanding = sess.get("dataset_understanding")
    schema_index = (sess.get("indexes") or {}).get("schema")
    if not understanding or not schema_index:
        return jsonify({"status": "error", "error": "Build the understanding card first."}), 200
    return jsonify(understanding_engine.generate_understanding_questions(understanding, schema_index))


@app.route("/api/understanding/generate", methods=["POST"])
def understanding_generate():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    question = body.get("question")
    if not store.session_exists(sid) or not question:
        return jsonify({"error": "session_id and question required"}), 400
    sess = store.get_session(sid)
    understanding = sess.get("dataset_understanding")
    schema_index = (sess.get("indexes") or {}).get("schema")
    if not understanding or not schema_index:
        return jsonify({"status": "error", "error": "Build the understanding card first."}), 200
    return jsonify(understanding_engine.answer_understanding_question(question, understanding, schema_index))


@app.route("/api/derive/plan", methods=["GET"])
def derive_plan():
    sid = request.args.get("session_id")
    if not sid or not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    dfs = _load_all_dfs(sdir)
    return jsonify(derive_contextual.get_plan(sess, dfs=dfs))


@app.route("/api/derive", methods=["POST"])
def derive():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404

    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    dfs = _load_all_dfs(sdir)
    # Never let a previous generic/time-series/expression/grouped derive
    # pass's own output (profile, numeric_summary, group_stats, ...) get
    # re-classified or re-derived as if it were fresh source data — that's
    # not idempotent and compounds into nonsense on a second derive call.
    source_dfs = derive_contextual._exclude_generic_derived(dfs)
    plan = derive_contextual.get_plan(sess, dfs=dfs)
    context = plan["context"]

    try:
        if context == "clinical_trial":
            result = _derive_clinical(sess, sdir)
        elif context == "lab_assay":
            result = _derive_lab(sdir)
        elif context == "plate_assay":
            result = _derive_plate(sdir)
        elif context == "time_series":
            result = _derive_time_series(sdir, source_dfs)
        elif context == "expression_matrix":
            result = _derive_expression(sdir, source_dfs)
        elif context == "grouped_comparison":
            result = _derive_grouped(sdir, source_dfs)
        else:
            result = _derive_generic(sdir)
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 200

    if result["status"] != "ok":
        return jsonify(result), 200

    if "provenance" in result:
        sess["derivation_meta"] = result["provenance"]

    datasets = {}
    for key, df in result["derived"].items():
        notebook_engine.save_state(sdir, key, df)
        datasets[key] = {"rows": len(df), "columns": list(df.columns)}

    sess.setdefault("derived_vars", [])
    sess["derived_vars"] = sorted(set(sess["derived_vars"]) | set(datasets.keys()))

    return jsonify({
        "status": "ok",
        "context": context,
        "context_label": plan["context_label"],
        "datasets": datasets,
    })


def _derive_clinical(sess, sdir):
    missing = [d for d in ["DM", "EX", "AE", "RS", "DS"] if d not in sess.get("domain_data", {})]
    if missing:
        return {"status": "incomplete", "missing_domains": missing}

    dm = notebook_engine.load_state(sdir, "dm")
    ex = notebook_engine.load_state(sdir, "ex")
    ae = notebook_engine.load_state(sdir, "ae")
    ds = notebook_engine.load_state(sdir, "ds")
    # DV (protocol deviations) is optional — collected whenever ingested, but
    # never blocks derivation the way the 5 required domains do.
    dv = notebook_engine.load_state(sdir, "dv") if "DV" in sess.get("domain_data", {}) else None

    if "AGE" in dm.columns:
        dm["AGE"] = pd.to_numeric(dm["AGE"], errors="coerce")
    if "ECOGBL" in dm.columns:
        dm["ECOGBL"] = pd.to_numeric(dm["ECOGBL"], errors="coerce")

    adsl  = derive_adam.derive_adsl(dm, ex, ds, dv=dv)
    adae  = derive_adam.derive_adae(ae, adsl)
    adtte = derive_adam.derive_adtte(adsl)
    derived = {"adsl": adsl, "adae": adae, "adtte": adtte}
    if dv is not None and not dv.empty:
        derived["addv"] = derive_adam.derive_addv(dv, adsl)

    provenance = {
        "recipe": "clinical_trial",
        "fired_because": (
            "All 5 SDTM domains present (DM, EX, AE, RS, DS)" +
            (" plus DV (protocol deviations)" if dv is not None and not dv.empty else "") + ". "
            "ADaM pipeline: ADSL from DM+EX+DS" + ("+DV" if dv is not None and not dv.empty else "") +
            ", ADAE from AE+ADSL, ADTTE (OS endpoint) from ADSL" +
            (", ADDV (deviation summary) from DV+ADSL" if dv is not None and not dv.empty else "") + "."
        ),
        "variable_origins": derive_adam.PIPELINE_PROVENANCE,
        "low_confidence": [],
    }
    return {"status": "ok", "derived": derived, "provenance": provenance}


def _derive_lab(sdir):
    lb = notebook_engine.load_state(sdir, "lb")
    lb_var = "lb"
    if lb is None:
        for v in notebook_engine.available_vars(sdir):
            df = notebook_engine.load_state(sdir, v)
            if df is not None and derive_contextual._col(df, ["LBORRES", "RESULT", "VALUE"]):
                lb = df
                lb_var = v
                break
    if lb is None:
        return {"status": "error", "error": "No lab data found in session."}

    lb_summary = derive_contextual.derive_lb_summary(lb)
    lb_flags   = derive_contextual.derive_lb_flags(lb)
    lb_shifts  = derive_contextual.derive_lb_shifts(lb)

    provenance = {
        "recipe": "lab_assay",
        "fired_because": (
            f"Lab data detected in '{lb_var}' (contains result/value column). "
            "Derived summary statistics, abnormality flags against reference range, and shift table."
        ),
        "variable_origins": {
            f"lb_summary.mean":    {"source": f"{lb_var}.LBORRES",                    "transform": "grouped mean by test + visit",                   "confidence": 1.00},
            f"lb_flags.HIGH_FLAG": {"source": f"{lb_var}.LBORRES vs LBORNRHI",        "transform": "value > upper limit of normal reference range",  "confidence": 0.95},
            f"lb_flags.LOW_FLAG":  {"source": f"{lb_var}.LBORRES vs LBORNRLO",        "transform": "value < lower limit of normal reference range",  "confidence": 0.95},
            f"lb_shifts.SHIFT":    {"source": f"{lb_var}.LBORRES + reference range",  "transform": "normal/high/low classification per visit",       "confidence": 0.90},
        },
        "low_confidence": [],
    }
    return {"status": "ok", "derived": {"lb_summary": lb_summary, "lb_flags": lb_flags, "lb_shifts": lb_shifts}, "provenance": provenance}


def _derive_plate(sdir):
    signal_df = compound_df = metadata_df = None
    signal_var = compound_var = metadata_var = None

    for v in notebook_engine.available_vars(sdir):
        df = notebook_engine.load_state(sdir, v)
        if df is None:
            continue
        cols_up = {c.upper() for c in df.columns}
        if "WELLID" in cols_up and "READOUT" in cols_up:
            numeric_frac = pd.to_numeric(df["READOUT"], errors="coerce").notna().mean()
            if numeric_frac > 0.7:
                if signal_df is None:
                    signal_df = df
                    signal_var = v
            else:
                if compound_df is None:
                    compound_df = df
                    compound_var = v
        elif "WELLID" not in cols_up and df.shape[0] < 50:
            metadata_df = df

    if signal_df is None:
        return {"status": "error", "error": "No plate signal data found — need a sheet with WELLID + numeric READOUT."}

    plate = signal_df.copy()
    plate["SIGNAL"] = pd.to_numeric(plate["READOUT"], errors="coerce")
    plate = plate.drop(columns=["READOUT"])

    if compound_df is not None:
        comp = compound_df[["WELLID", "READOUT"]].copy()

        def _parse_treatment(val):
            if pd.isna(val):
                return str(val), None
            s = str(val).strip()
            for sep in ("@", "_", " "):
                m = re.match(
                    r'^(.+?)' + re.escape(sep) + r'([\d.]+)\s*(?:um|µm|nm|mm|pm)?$',
                    s, re.IGNORECASE
                )
                if m:
                    return m.group(1).strip(), float(m.group(2))
            return s, None

        parsed = comp["READOUT"].apply(_parse_treatment)
        comp["COMPOUND"] = parsed.apply(lambda t: t[0])
        comp["CONCENTRATION"] = pd.to_numeric(parsed.apply(lambda t: t[1]), errors="coerce")

        _blank_names  = {"blank", "media", "background", "pbs"}
        _dmso_names   = {"dmso", "vehicle", "control", "neg", "untreated"}
        def _control_type(name):
            n = str(name).lower()
            if n in _blank_names:  return "blank"
            if n in _dmso_names:   return "dmso"
            return "treated"
        comp["CONTROL_TYPE"] = comp["COMPOUND"].apply(_control_type)

        comp = comp.drop(columns=["READOUT"])
        plate = plate.merge(comp, on="WELLID", how="left")

    if metadata_df is not None:
        rows_col = derive_contextual._col(metadata_df, ["PLATE_ROWS", "ROWS", "ROW_RANGE"])
        name_col = (derive_contextual._col(metadata_df, ["CELL_LINE_NAME", "CELL_LINE_ID", "SAMPLEID"]) or
                    next((c for c in metadata_df.columns if "name" in c.lower() or "id" in c.lower()), None))
        if rows_col and name_col and "ROW" in plate.columns:
            for _, row in metadata_df.iterrows():
                spec = str(row[rows_col]).strip()
                label = str(row[name_col])
                if "-" in spec:
                    a, b = spec.split("-", 1)
                    letters = [chr(r) for r in range(ord(a.strip().upper()), ord(b.strip().upper()) + 1)]
                else:
                    letters = [r.strip().upper() for r in spec.split(",")]
                plate.loc[plate["ROW"].isin(letters), "CELL_LINE"] = label

    if "CONTROL_TYPE" in plate.columns:
        blank_mask = plate["CONTROL_TYPE"] == "blank"
        dmso_mask  = plate["CONTROL_TYPE"] == "dmso"
        blank_mean = plate.loc[blank_mask, "SIGNAL"].mean() if blank_mask.any() else 0.0
        if pd.isna(blank_mean):
            blank_mean = 0.0

        group_col = "CELL_LINE" if "CELL_LINE" in plate.columns else None
        groups = plate.groupby(group_col) if group_col else [(None, plate)]
        normed = []
        for _, grp in groups:
            grp = grp.copy()
            dmso_mean = grp.loc[dmso_mask.loc[grp.index], "SIGNAL"].mean()
            if pd.isna(dmso_mean) or dmso_mean == blank_mean:
                dmso_mean = grp["SIGNAL"].max()
            denom = dmso_mean - blank_mean
            grp["VIABILITY_PCT"] = ((grp["SIGNAL"] - blank_mean) / denom * 100).round(2) if denom else 0.0
            normed.append(grp)
        plate = pd.concat(normed).sort_index()

    plate_qc = derive_contextual.derive_plate_qc(plate.rename(columns={"SIGNAL": "READOUT"}, errors="ignore"))

    viability_col = "VIABILITY_PCT" if "VIABILITY_PCT" in plate.columns else "SIGNAL"
    treat_mask = plate["CONTROL_TYPE"] == "treated" if "CONTROL_TYPE" in plate.columns else pd.Series(True, index=plate.index)
    treat = plate[treat_mask]
    if "CONCENTRATION" in plate.columns and treat_mask.any():
        group_cols = ["CELL_LINE", "CONCENTRATION"] if "CELL_LINE" in treat.columns else ["CONCENTRATION"]
        dose_response = (
            treat.groupby(group_cols)[viability_col]
            .agg(N="count", mean_viability="mean", sd="std")
            .round(2).reset_index()
        )
    else:
        dose_response = derive_contextual.derive_dose_response(plate)

    origins = {}
    origins["plate_assay.SIGNAL"] = {
        "source": f"{signal_var}.READOUT",
        "transform": "coerced to numeric via pd.to_numeric",
        "confidence": 1.00,
    }
    if compound_df is not None:
        origins["plate_assay.COMPOUND"] = {
            "source": f"{compound_var}.READOUT",
            "transform": "treatment string parsed: name before separator (@, _, space)",
            "confidence": 0.90,
        }
        origins["plate_assay.CONCENTRATION"] = {
            "source": f"{compound_var}.READOUT",
            "transform": "treatment string parsed: numeric suffix, unit stripped (µM)",
            "confidence": 0.90,
        }
        origins["plate_assay.CONTROL_TYPE"] = {
            "source": f"{compound_var}.COMPOUND",
            "transform": "rule-based classification: blank / dmso / treated",
            "confidence": 0.95,
        }
    if "VIABILITY_PCT" in plate.columns:
        origins["plate_assay.VIABILITY_PCT"] = {
            "source": "plate_assay.SIGNAL",
            "transform": "(signal − blank_mean) / (dmso_mean − blank_mean) × 100",
            "confidence": 0.85,
        }
    origins["dose_response.CONCENTRATION"] = {
        "source": "plate_assay.CONCENTRATION",
        "transform": "groupby key (treated wells only, per CELL_LINE)",
        "confidence": 0.90,
    }
    origins["dose_response.mean_viability"] = {
        "source": "plate_assay.VIABILITY_PCT",
        "transform": "mean per concentration per cell line",
        "confidence": 0.85,
    }

    sheet_parts = [f"{signal_var} (numeric signal)"]
    if compound_df is not None:
        sheet_parts.append(f"{compound_var} (treatment strings → compound + concentration)")
    provenance = {
        "recipe": "plate_assay",
        "fired_because": (
            f"Detected {len(sheet_parts)} plate-format sheet(s): {' + '.join(sheet_parts)}. "
            "Sheets joined on WELLID. Treatment strings parsed into COMPOUND + CONCENTRATION columns."
        ),
        "variable_origins": origins,
        "low_confidence": [k for k, v in origins.items() if v["confidence"] < 0.90],
    }

    return {
        "status": "ok",
        "derived": {"plate_assay": plate, "plate_qc": plate_qc, "dose_response": dose_response},
        "provenance": provenance,
    }


def _derive_time_series(sdir, dfs):
    if not dfs:
        return {"status": "error", "error": "No data found in session."}
    primary_var, primary_df = next(iter(dfs.items()))

    time_col = next(
        (c for c in primary_df.columns
         if any(kw in c.lower() for kw in derive_contextual._TIME_KEYWORDS)),
        None,
    )
    if time_col is None:
        return _derive_generic(sdir)

    time_index    = derive_contextual.derive_time_index(primary_df, time_col)
    trend_summary = derive_contextual.derive_trend_summary(primary_df, time_col)

    n_vars = len(primary_df.select_dtypes(include="number").columns)
    provenance = {
        "recipe": "time_series",
        "fired_because": (
            f"Column '{time_col}' in '{primary_var}' identified as a temporal index "
            f"({len(primary_df)} observations, {n_vars} numeric variable(s)). "
            "Derived time-ordered index with inter-observation delta and per-variable linear trend analysis."
        ),
        "variable_origins": {
            "time_index.TIME_DELTA": {
                "source": f"{primary_var}.{time_col}",
                "transform": "consecutive difference after chronological sort",
                "confidence": 0.95,
            },
            "trend_summary.slope": {
                "source": f"{primary_var} numeric columns vs '{time_col}'",
                "transform": "numpy.polyfit(deg=1) — slope and R² per variable",
                "confidence": 0.90,
            },
        },
        "low_confidence": [],
    }
    return {
        "status": "ok",
        "derived": {"time_index": time_index, "trend_summary": trend_summary},
        "provenance": provenance,
    }


def _derive_expression(sdir, dfs):
    if not dfs:
        return {"status": "error", "error": "No data found in session."}
    primary_var, primary_df = next(iter(dfs.items()))

    num_cols = primary_df.select_dtypes(include="number").columns.tolist()
    n_features = len(primary_df)
    n_samples = len(num_cols)

    feature_variance = derive_contextual.derive_feature_variance(primary_df)
    sample_stats     = derive_contextual.derive_sample_stats(primary_df)
    top_variable     = derive_contextual.derive_top_variable(primary_df, n=50)

    provenance = {
        "recipe": "expression_matrix",
        "fired_because": (
            f"Wide numeric matrix detected in '{primary_var}': {n_features} feature rows × "
            f"{n_samples} sample columns ({round(n_samples / len(primary_df.columns) * 100)}% numeric). "
            "Derived per-feature variance ranking, per-sample detection statistics, and top-50 variable features."
        ),
        "variable_origins": {
            "feature_variance.variance": {
                "source": f"{primary_var} (all {n_samples} sample columns)",
                "transform": "pandas DataFrame.var(axis=1) per feature row",
                "confidence": 1.00,
            },
            "sample_stats.pct_detected": {
                "source": f"{primary_var} (all {n_features} feature rows)",
                "transform": "fraction of values > 0 per sample column",
                "confidence": 1.00,
            },
            "top_variable": {
                "source": "feature_variance",
                "transform": f"head(50) by descending variance from {n_features} features",
                "confidence": 1.00,
            },
        },
        "low_confidence": [],
    }
    return {
        "status": "ok",
        "derived": {"feature_variance": feature_variance, "sample_stats": sample_stats, "top_variable": top_variable},
        "provenance": provenance,
    }


def _derive_grouped(sdir, dfs):
    if not dfs:
        return {"status": "error", "error": "No data found in session."}
    primary_var, primary_df = next(iter(dfs.items()))

    group_col, num_cols = derive_contextual._pick_group_cols(primary_df)
    if not group_col:
        return _derive_generic(sdir)

    n_groups = primary_df[group_col].nunique()

    group_stats   = derive_contextual.derive_group_stats(primary_df)
    effect_sizes  = derive_contextual.derive_effect_sizes(primary_df)
    anova_summary = derive_contextual.derive_anova(primary_df)

    provenance = {
        "recipe": "grouped_comparison",
        "fired_because": (
            f"Group column '{group_col}' detected in '{primary_var}' ({n_groups} groups, "
            f"{len(num_cols)} numeric measurement(s)). Derived group statistics, Cohen's d effect sizes, "
            "and one-way ANOVA significance tests per measurement."
        ),
        "variable_origins": {
            "group_stats.mean": {
                "source": f"{primary_var}.{group_col} + numeric columns",
                "transform": "groupby mean / std / median / N",
                "confidence": 1.00,
            },
            "effect_sizes.cohens_d": {
                "source": f"{primary_var} — pairwise group combinations",
                "transform": "pooled-SD Cohen's d for each group pair × measurement",
                "confidence": 0.90,
            },
            "anova_summary.p_value": {
                "source": f"{primary_var} — all {n_groups} groups per measurement",
                "transform": "scipy.stats.f_oneway — one-way ANOVA",
                "confidence": 0.90,
            },
        },
        "low_confidence": [],
    }
    return {
        "status": "ok",
        "derived": {"group_stats": group_stats, "effect_sizes": effect_sizes, "anova_summary": anova_summary},
        "provenance": provenance,
    }


def _derive_generic(sdir):
    all_vars = notebook_engine.available_vars(sdir)
    dfs = {v: notebook_engine.load_state(sdir, v) for v in all_vars}
    dfs = {v: df for v, df in dfs.items() if df is not None}
    dfs = derive_contextual._exclude_generic_derived(dfs)
    if not dfs:
        return {"status": "error", "error": "No data found in session to derive from."}

    profile         = derive_contextual.derive_profile(dfs)
    numeric_summary = derive_contextual.derive_numeric_summary(dfs)

    src_list = ", ".join(dfs.keys())
    provenance = {
        "recipe": "generic",
        "fired_because": (
            f"No specific scientific domain detected. Generated statistical profile and "
            f"numeric summary from all available variables ({src_list})."
        ),
        "variable_origins": {
            "profile":         {"source": src_list, "transform": "shape, dtypes, missing count, unique count per variable", "confidence": 1.00},
            "numeric_summary": {"source": src_list, "transform": "min / max / mean / std / quartiles for all numeric columns",  "confidence": 1.00},
        },
        "low_confidence": [],
    }
    return {"status": "ok", "derived": {"profile": profile, "numeric_summary": numeric_summary}, "provenance": provenance}


def _current_context(sess) -> str:
    try:
        return derive_contextual.get_plan(sess).get("context", "generic")
    except Exception:
        return "generic"


@app.route("/api/index/build", methods=["POST"])
def index_build():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404

    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    dfs = _load_all_dfs(sdir)
    if not dfs:
        return jsonify({"status": "error", "error": "No derived data yet — run derivation first."}), 200

    context = _current_context(sess)
    built = indexer.build_indexes(dfs, context, existing_dashboards=sess.get("dashboards", []), sid=sid)
    sess["dataset_understanding"] = built["understanding"]
    sess["indexes"] = built["indexes"]

    canonical_vars = {e.get("var_name", d.lower()) for d, e in sess.get("domain_source", {}).items()}
    canonical_vars |= set(sess.get("derived_vars", []))
    version = dataset_registry.register_version(sid, dfs, canonical_vars=canonical_vars or None,
                                                 understanding=built["understanding"])
    sess["dataset_fingerprint"] = version["fingerprint"]
    sess["dataset_version_id"] = version["id"]
    sess["_understanding_built_for_fingerprint"] = version["fingerprint"]

    return jsonify({
        "status": "ok",
        "context": context,
        "understanding": built["understanding"],
        "indexes": {
            "dataset": built["indexes"]["dataset"],
            "entity": built["indexes"]["entity"],
            "metric": built["indexes"]["metric"],
            "cohort": built["indexes"]["cohort"],
            "narrative": built["indexes"]["narrative"],
            "finding": built["indexes"]["finding"],
            "schema_var_count": {v: len(cols) for v, cols in built["indexes"]["schema"].items()},
        },
    })


@app.route("/api/index/understanding", methods=["GET"])
def index_understanding():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    if not sess.get("dataset_understanding"):
        return jsonify({"status": "not_built"}), 200
    return jsonify({
        "status": "ok",
        "understanding": sess["dataset_understanding"],
        "indexes": {
            "dataset": sess["indexes"]["dataset"],
            "entity": sess["indexes"]["entity"],
            "metric": sess["indexes"]["metric"],
            "cohort": sess["indexes"]["cohort"],
        },
    })


def _ensure_indexes(sid, sess, sdir):
    """Returns (understanding, schema_index, dfs, error_response_or_None).

    Registers/looks up the dataset version for the current dataframes on
    every call (cheap — a content hash, no LLM) and rebuilds the classifier
    output whenever that fingerprint doesn't match what understanding was
    last built from. Before dataset_registry existed, understanding was only
    ever built once per session and silently went stale after any later
    upload/derive — this is the fix for that, not just a cache.

    Act 2/3 (triple-trial) sessions skip all of this — /api/index/build-triple
    already built the combined understanding/schema explicitly, keyed on
    suffixed table names the single-trial classifier below doesn't know
    about, and dataset-version caching isn't meaningful across 3 datasets
    at once.
    """
    if sess.get("act", 1) >= 2:
        dfs = _load_triple_dfs(sdir)
        if not dfs:
            return None, None, None, ({"status": "error", "error": "No triple-trial data derived yet."}, 200)
        if not sess.get("dataset_understanding"):
            return None, None, None, ({"status": "error", "error": "Build contextual understanding first."}, 200)
        return sess["dataset_understanding"], sess["indexes"]["schema"], dfs, None

    dfs = _load_all_dfs(sdir)
    if not dfs:
        return None, None, None, ({"status": "error", "error": "No derived data yet — run derivation first."}, 200)

    canonical_vars = {e.get("var_name", d.lower()) for d, e in sess.get("domain_source", {}).items()}
    canonical_vars |= set(sess.get("derived_vars", []))
    version = dataset_registry.register_version(sid, dfs, canonical_vars=canonical_vars or None,
                                                 understanding=sess.get("dataset_understanding"))
    sess["dataset_fingerprint"] = version["fingerprint"]
    sess["dataset_version_id"] = version["id"]

    stale = version["fingerprint"] != (sess.get("_understanding_built_for_fingerprint"))
    if not sess.get("dataset_understanding") or stale:
        context = _current_context(sess)
        built = indexer.build_indexes(dfs, context, existing_dashboards=sess.get("dashboards", []))
        sess["dataset_understanding"] = built["understanding"]
        sess["indexes"] = built["indexes"]
        sess["_understanding_built_for_fingerprint"] = version["fingerprint"]
        db.update_dataset_version_understanding(version["id"], built["understanding"])
    return sess["dataset_understanding"], sess["indexes"]["schema"], dfs, None


@app.route("/api/dashboards/candidates", methods=["GET"])
def dashboards_candidates():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    understanding, schema_index, dfs, error = _ensure_indexes(sid, sess, sdir)
    if error:
        return jsonify(error[0]), error[1]
    return jsonify(dashboard_engine.list_candidates(dfs, understanding, schema_index, compare_mode=sess.get("act", 1) >= 2))


@app.route("/api/dashboards/generate", methods=["POST"])
def dashboards_generate():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    mode = body.get("mode", "autopilot")
    question = body.get("question")

    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    understanding, schema_index, dfs, error = _ensure_indexes(sid, sess, sdir)
    if error:
        return jsonify(error[0]), error[1]
    fingerprint = sess.get("dataset_fingerprint")
    version_id = sess.get("dataset_version_id")

    if mode == "copilot":
        if not question:
            return jsonify({"status": "error", "error": "question required for copilot mode"}), 200
        result = dashboard_engine.generate_copilot(sess, sdir, understanding, schema_index, question,
                                                    dataset_fingerprint=fingerprint, sid=sid, dataset_version_id=version_id)
    else:
        result = dashboard_engine.generate_autopilot(sess, sdir, dfs, understanding, schema_index,
                                                       dataset_fingerprint=fingerprint, sid=sid, dataset_version_id=version_id)

    if sess.get("indexes"):
        sess["indexes"]["dashboard"] = [
            {"dashboard_id": d["dashboard_id"], "question": d["question"]} for d in sess["dashboards"]
        ]

    return jsonify(result)


@app.route("/api/dashboards/followups", methods=["POST"])
def dashboards_followups():
    """Tree-shaped investigation: given a question already answered (and what
    was actually found), ask the Hypothesis Agent what to look at next. This
    is the mechanism that replaces a static "try asking" suggestion list —
    every suggestion past the first pass is grounded in a real prior result,
    not a blind schema scan."""
    body = request.get_json(force=True)
    sid = body.get("session_id")
    question = body.get("question")
    stats = body.get("stats") or {}
    chart_type = body.get("chart_type")
    n = body.get("n", 1)

    if not store.session_exists(sid) or not question:
        return jsonify({"error": "session_id and question required"}), 400
    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    understanding, schema_index, dfs, error = _ensure_indexes(sid, sess, sdir)
    if error:
        return jsonify(error[0]), error[1]

    return jsonify(dashboard_engine.generate_followups(understanding, schema_index, dfs, question, stats, chart_type, n=n))


@app.route("/api/analysis/candidates", methods=["GET"])
def analysis_candidates():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    understanding = sess.get("dataset_understanding") or {}
    notebook_results = sess.get("dashboards", [])
    return jsonify(synthesis_engine.generate_synthesis_questions(understanding, notebook_results))


@app.route("/api/analysis/generate", methods=["POST"])
def analysis_generate():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    question = body.get("question")
    relevant_indices = body.get("relevant_result_indices") or []
    if not store.session_exists(sid) or not question:
        return jsonify({"error": "session_id and question required"}), 400
    sess = store.get_session(sid)
    understanding = sess.get("dataset_understanding") or {}
    notebook_results = sess.get("dashboards", [])
    result = synthesis_engine.answer_synthesis_question(
        sid, sess.get("dataset_version_id"), question, notebook_results, relevant_indices, understanding,
    )
    return jsonify(result)


@app.route("/api/dashboards", methods=["GET"])
def dashboards_list():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    catalog = [
        {k: v for k, v in d.items() if k != "code"}
        for d in sess.get("dashboards", [])
    ]
    return jsonify({"dashboards": catalog})


@app.route("/api/dashboards/<did>", methods=["GET"])
def dashboards_get(did):
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    artifact = next((d for d in sess.get("dashboards", []) if d["dashboard_id"] == did), None)
    if artifact is None:
        return jsonify({"error": "unknown dashboard"}), 404
    return jsonify(artifact)


@app.route("/api/dashboards/promote", methods=["POST"])
def dashboards_promote():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    cell_id = body.get("cell_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    cell = next((c for c in sess.get("canvas_cells", []) if c["id"] == cell_id), None)
    if cell is None:
        return jsonify({"status": "error", "error": "unknown cell_id"}), 200
    result = dashboard_engine.promote_cell(sess, cell)
    return jsonify(result)


@app.route("/api/session/<sid>/provenance", methods=["GET"])
def session_provenance(sid):
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)

    domain_source   = sess.get("domain_source", {})
    derivation_meta = sess.get("derivation_meta", {})
    files           = sess.get("files", {})

    raw_origins:    dict = {}
    held_for_review: list = []

    for domain, ds_meta in domain_source.items():
        var_name = domain.lower()
        filename = ds_meta.get("filename", "")
        sheet    = ds_meta.get("sheet_name")

        file_result = files.get(filename, {})
        if file_result.get("detected_format") == "xlsx" and sheet:
            mapping = next(
                (s.get("mapping", []) for s in file_result.get("sheets", [])
                 if s.get("sheet_name") == sheet),
                [],
            )
        else:
            mapping = file_result.get("mapping", [])

        for entry in mapping:
            action     = entry.get("action", "")
            source_col = entry.get("col", "")
            sdtm_col   = entry.get("top_match", source_col)
            confidence = entry.get("confidence", 0.0)

            if action not in ("AUTO_MAP", "SURFACE_TO_USER"):
                continue

            ref = f"{var_name}.{sdtm_col}"
            raw_origins[ref] = {
                "source_file": filename + (f":{sheet}" if sheet else ""),
                "source_col":  source_col,
                "confidence":  round(confidence, 2),
                "action":      action,
            }

            if action == "SURFACE_TO_USER":
                held_for_review.append({
                    "var":        var_name,
                    "col":        sdtm_col,
                    "source_col": source_col,
                    "mapped_to":  sdtm_col,
                    "confidence": round(confidence, 2),
                })

    return jsonify({
        "domain_source":   domain_source,
        "derivation_meta": derivation_meta,
        "raw_origins":     raw_origins,
        "held_for_review": held_for_review,
    })


@app.route("/api/notebook/run", methods=["POST"])
def notebook_run():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    code = body.get("code", "")

    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    if not code.strip():
        return jsonify({"status": "error", "error": "empty cell"}), 200

    sdir = _session_dir(sid)
    result = notebook_engine.run_cell(sdir, code)

    sess = store.get_session(sid)
    cell_record = {"id": len(sess["canvas_cells"]) + 1, "code": code, "result": result}
    sess["canvas_cells"].append(cell_record)

    return jsonify(cell_record)


@app.route("/api/session/<sid>/notebook", methods=["GET"])
def notebook_list(sid):
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    return jsonify({"cells": sess["canvas_cells"]})


def _detect_quality_issues(df, var_name: str) -> list:
    issues = []

    dup = int(df.duplicated().sum())
    if dup:
        issues.append({"var": var_name, "col": None, "type": "duplicates",
                        "description": f"{dup} duplicate row{'s' if dup>1 else ''}",
                        "fix": "drop_duplicates", "fix_label": "Drop duplicate rows",
                        "severity": "medium"})

    for col in df.columns:
        s = df[col]
        null_count = int(s.isna().sum())
        pct = null_count / max(len(df), 1) * 100

        if null_count:
            is_numeric = pd.api.types.is_numeric_dtype(s)
            fill_label = (f"Fill with mean ({s.mean():.3g})" if is_numeric else "Fill with 'UNKNOWN'")
            issues.append({"var": var_name, "col": col, "type": "missing",
                            "description": f"{null_count} null values ({pct:.1f}%)",
                            "fix": "fill_mean" if is_numeric else "fill_unknown",
                            "fix_label": fill_label,
                            "severity": "high" if pct > 20 else "medium" if pct > 5 else "low"})

        if pd.api.types.is_numeric_dtype(s) and s.notna().sum() > 4:
            mu, sigma = float(s.mean()), float(s.std())
            if sigma > 0:
                n_out = int(((s - mu).abs() > 3 * sigma).sum())
                if n_out:
                    issues.append({"var": var_name, "col": col, "type": "outlier",
                                    "description": f"{n_out} value{'s' if n_out>1 else ''} beyond 3σ (mean={mu:.3g}, σ={sigma:.3g})",
                                    "fix": "cap_3sigma", "fix_label": "Cap at ±3σ",
                                    "severity": "medium"})

        if s.dtype == object and 1 < s.nunique() <= 30:
            vals = s.dropna().astype(str)
            if vals.str.lower().nunique() < vals.nunique():
                sample = sorted(vals.unique())[:3]
                issues.append({"var": var_name, "col": col, "type": "mixed_case",
                                "description": f"Mixed case — e.g. {sample}",
                                "fix": "normalize_upper", "fix_label": "Normalise to uppercase",
                                "severity": "low"})

    return issues


def _apply_fix(df: pd.DataFrame, fix: dict) -> pd.DataFrame:
    col = fix.get("col")
    fix_type = fix.get("fix")
    if fix_type == "drop_duplicates":
        return df.drop_duplicates()
    if col not in df.columns:
        return df
    df = df.copy()
    if fix_type == "fill_mean":
        df[col] = df[col].fillna(df[col].mean())
    elif fix_type == "fill_unknown":
        df[col] = df[col].fillna("UNKNOWN")
    elif fix_type == "cap_3sigma":
        mu, sigma = df[col].mean(), df[col].std()
        df[col] = df[col].clip(mu - 3*sigma, mu + 3*sigma)
    elif fix_type == "normalize_upper":
        df[col] = df[col].str.upper()
    return df


@app.route("/api/quality/check", methods=["GET"])
def quality_check():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sdir = _session_dir(sid)
    all_issues = []

    for v in notebook_engine.available_vars(sdir):
        df = notebook_engine.load_state(sdir, v)
        if df is not None:
            all_issues.extend(_detect_quality_issues(df, v))

    clinical_issues = _validate.run_clinical_quality(
        sdir,
        notebook_engine.available_vars,
        notebook_engine.load_state,
    )

    clinical_covered = {(i.get("var"), i.get("col")) for i in clinical_issues if i.get("col")}
    filtered_generic = [
        i for i in all_issues
        if not (i["type"] == "missing" and (i["var"], i["col"]) in clinical_covered)
    ]
    all_issues = filtered_generic + clinical_issues

    return jsonify({"issues": all_issues})


@app.route("/api/quality/apply", methods=["POST"])
def quality_apply():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    fixes = body.get("fixes", [])
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sdir = _session_dir(sid)
    by_var: dict[str, list] = {}
    for fix in fixes:
        by_var.setdefault(fix["var"], []).append(fix)
    for var_name, var_fixes in by_var.items():
        df = notebook_engine.load_state(sdir, var_name)
        if df is None:
            continue
        for fix in var_fixes:
            df = _apply_fix(df, fix)
        notebook_engine.save_state(sdir, var_name, df)
    return jsonify({"status": "ok", "applied": len(fixes)})


@app.route("/api/export", methods=["GET"])
def export_xlsx():
    import io as _io
    from datetime import datetime
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    sdir = _session_dir(sid)
    sess = store.get_session(sid)
    available = notebook_engine.available_vars(sdir)

    HEADER_FILL  = PatternFill("solid", fgColor="1A3A5C")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    ALT_FILL     = PatternFill("solid", fgColor="EEF3F8")
    NORMAL_FONT  = Font(size=10, name="Calibri")
    TITLE_FONT   = Font(bold=True, size=13, name="Calibri")
    META_FONT    = Font(size=10, name="Calibri", color="555555")
    thin         = Side(style="thin", color="CCCCCC")
    BORDER       = Border(bottom=thin)

    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Summary"
    ws0.column_dimensions["A"].width = 24
    ws0.column_dimensions["B"].width = 48

    try:
        plan = derive_contextual.get_plan(sess)
        ctx_label = plan.get("context_label", "General tabular data")
    except Exception:
        ctx_label = "General tabular data"

    user_info = sess.get("user_info", {})
    meta_rows = [
        ("PROBE EXPORT", None),
        (None, None),
        ("Name",         user_info.get("name", "")),
        ("Organisation", user_info.get("organization", "")),
        ("Email",        user_info.get("email", "")),
        ("Project",      user_info.get("project", "")),
        (None, None),
        ("Session",      sid),
        ("Exported",     datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Data context", ctx_label),
        ("Variables",    ", ".join(available)),
    ]
    for r, (label, value) in enumerate(meta_rows, 1):
        if label == "PROBE EXPORT":
            cell = ws0.cell(row=r, column=1, value=label)
            cell.font = TITLE_FONT
        elif label:
            ws0.cell(row=r, column=1, value=label).font = Font(bold=True, size=10, name="Calibri", color="1A3A5C")
            ws0.cell(row=r, column=2, value=value).font = META_FONT

    PREFERRED_ORDER = ["adsl","adae","adtte","plate_assay","dose_response","plate_qc",
                       "lb_summary","lb_flags","lb_shifts","profile","numeric_summary"]
    ordered = [v for v in PREFERRED_ORDER if v in available] + [v for v in available if v not in PREFERRED_ORDER]

    for var in ordered:
        df = notebook_engine.load_state(sdir, var)
        if df is None or df.empty:
            continue

        ws = wb.create_sheet(title=var[:31])
        ws.freeze_panes = "A2"

        for ci, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=ci, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        for ri, (_, row) in enumerate(df.iterrows(), 2):
            fill = ALT_FILL if ri % 2 == 0 else None
            for ci, val in enumerate(row, 1):
                v = None if pd.isna(val) else val
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = NORMAL_FONT
                cell.border = BORDER
                if fill:
                    cell.fill = fill
                if pd.api.types.is_float_dtype(df.iloc[:, ci-1]):
                    cell.number_format = "0.0000"

        for ci, col_name in enumerate(df.columns, 1):
            sample_vals = df.iloc[:, ci-1].astype(str).head(50)
            max_len = max(len(str(col_name)), sample_vals.str.len().max() if not sample_vals.empty else 8)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 10), 40)

        ws.row_dimensions[1].height = 18

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(buf, as_attachment=True,
                     download_name=f"probe_export_{sid[:8]}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/reports", methods=["GET"])
def reports_list():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    return jsonify({"status": "ok", "reports": reports.list_for_session(sid)})


@app.route("/api/reports/generate", methods=["POST"])
def reports_generate():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    source_type = body.get("source_type")
    source_id = body.get("source_id")
    if not store.session_exists(sid) or not source_type or not source_id:
        return jsonify({"error": "session_id, source_type, source_id required"}), 400
    result = reports.generate_pdf(sid, source_type, source_id)
    return jsonify(result), (200 if result["status"] == "ok" else 400)


@app.route("/api/reports/<rid>/download", methods=["GET"])
def reports_download(rid):
    report = db.get_report(rid)
    if report is None or not os.path.exists(report["file_path"]):
        return jsonify({"error": "unknown report"}), 404
    from flask import send_file
    return send_file(report["file_path"], as_attachment=True,
                      download_name=f"{report['name'][:50].replace('/', '-')}.pdf", mimetype="application/pdf")


@app.route("/api/notebook/vars", methods=["GET"])
def notebook_vars():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sdir = _session_dir(sid)
    sess = store.get_session(sid)
    available = notebook_engine.available_vars(sdir)
    try:
        plan = derive_contextual.get_plan(sess)
        context = plan.get("context", "generic")
        context_label = plan.get("context_label", "General tabular data")
    except Exception:
        context = "generic"
        context_label = "General tabular data"
    return jsonify({"vars": available, "context": context, "context_label": context_label})


@app.route("/api/notebook/generate", methods=["POST"])
def notebook_generate():
    import codegen

    body = request.get_json(force=True)
    sid = body.get("session_id")
    request_text = body.get("text", "").strip()

    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    if not request_text:
        return jsonify({"status": "error", "error": "empty request"}), 200

    sdir = _session_dir(sid)
    available = notebook_engine.available_vars(sdir)
    if not available:
        return jsonify({"status": "error", "error": "No derived data yet — run derivation first."}), 200

    sess = store.get_session(sid)
    gen_result = codegen.generate_code(sdir, available, request_text, sess=sess)
    if gen_result["status"] != "ok":
        err = gen_result.get("error", "")
        if "disallowed pattern" in err and any(p in err for p in ["open(", "import os", "import sys"]):
            reason = "cannot_write_files"
        elif "No derived data" in err or "No dataframes" in err:
            reason = "no_data"
        elif "ANTHROPIC_API_KEY" in err:
            reason = "api_key"
        else:
            reason = "general"
        return jsonify({"status": "error", "reason": reason}), 200

    code = gen_result["code"]

    dfs = _load_all_dfs(sdir)
    schema_check = codegen.verify_against_schema(code, dfs)

    exec_result = notebook_engine.run_cell(sdir, code)

    sess = store.get_session(sid)
    cell_record = {
        "id": len(sess["canvas_cells"]) + 1,
        "code": code,
        "generated_from": request_text,
        "schema_check": schema_check,
        "result": exec_result,
    }
    sess["canvas_cells"].append(cell_record)

    return jsonify(cell_record)


@app.route("/api/oracle/resolve", methods=["POST"])
def oracle_resolve():
    body = request.get_json(force=True)
    oracle_type = body.get("oracle_type")
    population_args = body.get("population_args", {})
    force = bool(body.get("force", False))
    if not oracle_type:
        return jsonify({"status": "error", "error": "oracle_type required"}), 200
    result = oracle_engine.resolve_oracle(oracle_type, population_args, force=force)
    return jsonify(result)


@app.route("/api/oracle/types", methods=["GET"])
def oracle_types():
    return jsonify({"status": "ok", "types": oracle_engine.ORACLE_TYPES})


@app.route("/api/oracle/<oid>", methods=["GET"])
def oracle_get(oid):
    inst = oracle_engine.get_instance(oid)
    if inst is None:
        return jsonify({"error": "unknown oracle instance"}), 404
    return jsonify({"status": "ok", "instance": inst})


@app.route("/api/oracle/<oid>/pin", methods=["POST"])
def oracle_pin(oid):
    return jsonify(oracle_engine.pin(oid))


@app.route("/api/oracle/<oid>/drift", methods=["POST"])
def oracle_drift(oid):
    body = request.get_json(silent=True) or {}
    return jsonify(oracle_engine.inject_drift_source(oid, body.get("source")))


@app.route("/api/oracle/<oid>/reset", methods=["POST"])
def oracle_reset(oid):
    return jsonify(oracle_engine.reset(oid))


@app.route("/api/narratives/generate", methods=["POST"])
def narratives_generate():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    sdir = _session_dir(sid)
    understanding, schema_index, dfs, error = _ensure_indexes(sid, sess, sdir)
    if error:
        return jsonify(error[0]), error[1]

    result = narrative_engine.generate_narrative(sid, sess, sdir, dfs, understanding, schema_index,
                                                  dataset_fingerprint=sess.get("dataset_fingerprint"),
                                                  dataset_version_id=sess.get("dataset_version_id"),
                                                  enable_oracles=bool(sess.get("oracles_enabled", False)))
    return jsonify(result)


@app.route("/api/act/enable-oracles", methods=["POST"])
def act_enable_oracles():
    """The Act 2 -> Act 3 transition: from here on Evidence benchmarks
    against outside sources and Narratives cite them — nothing else about
    the session (trials, notebook results, analysis) changes or re-runs."""
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    sess["oracles_enabled"] = True
    sess["act"] = 3
    return jsonify({"status": "ok", "act": 3, "oracles_enabled": True})


@app.route("/api/narratives", methods=["GET"])
def narratives_list():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    rows = db.list_narratives(sid, include_synthetic=False)
    rows.sort(key=lambda n: n.get("created_at", 0), reverse=True)
    return jsonify({"status": "ok", "narratives": rows})


@app.route("/api/datasets/<sid>/versions", methods=["GET"])
def dataset_versions(sid):
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    versions = dataset_registry.list_versions(sid)
    stats = db.cache_stats(versions[-1]["fingerprint"]) if versions else {"cached_artifacts": 0, "cache_hits": 0}
    return jsonify({"status": "ok", "versions": versions, "current_cache_stats": stats})


def _quality_tier(avg_missing_pct: float) -> list:
    if avg_missing_pct > 20:
        return ["red", "Missingness flagged"]
    if avg_missing_pct > 5:
        return ["amber", f"{avg_missing_pct:.0f}% missing"]
    return ["green", "Clean"]


@app.route("/api/datasets/<sid>/tables", methods=["GET"])
def dataset_tables(sid):
    """Real per-table catalog — every dataset here is honestly one session's
    set of canonical tables (raw domains + derived ADaM/etc outputs), not a
    multi-file upload catalog with independent version histories. Built from
    the same dataset-version manifest and schema index the rest of the
    pipeline already computes, not new data."""
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    version = dataset_registry.get_current(sid)
    if version is None:
        return jsonify({"status": "ok", "tables": []})

    schema = (sess.get("indexes") or {}).get("schema", {})
    tables = []
    for var, meta in version["manifest"].items():
        cols_meta = schema.get(var, {})
        missing_vals = [c["missing_pct"] for c in cols_meta.values()] if cols_meta else []
        avg_missing = sum(missing_vals) / len(missing_vals) if missing_vals else 0.0
        tables.append({
            "name": var, "rows": meta["rows"], "cols": meta["cols"],
            "dtypes": meta["dtypes"], "avg_missing_pct": round(avg_missing, 1),
            "quality": _quality_tier(avg_missing),
            "dataset_version": version["version"], "dataset_version_id": version["id"],
        })
    tables.sort(key=lambda t: t["name"])
    return jsonify({"status": "ok", "tables": tables, "dataset_version": version["version"]})


@app.route("/api/narratives/<nid>", methods=["GET"])
def narratives_get(nid):
    narrative = db.get_narrative(nid)
    if narrative is None:
        return jsonify({"error": "unknown narrative"}), 404
    return jsonify({"status": "ok", "narrative": narrative})


@app.route("/api/corpus", methods=["GET"])
def corpus_list():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    corpus_store.ensure_filler(sid, understanding=sess.get("dataset_understanding"))
    rows = corpus_store.list_corpus(sid, status=request.args.get("status"), q=request.args.get("q"))
    return jsonify({"status": "ok", "narratives": rows})


@app.route("/api/registry", methods=["GET"])
def registry_list():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    rows = hypothesis_registry.list_for_session(sid)
    return jsonify({"status": "ok", "rows": rows})


@app.route("/api/registry/publish", methods=["POST"])
def registry_publish():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    claim = body.get("claim")
    if not store.session_exists(sid) or not claim:
        return jsonify({"error": "session_id and claim required"}), 400
    result = hypothesis_registry.publish(sid, claim, narrative_id=body.get("narrative_id"),
                                          dag=body.get("dag"), q_value=body.get("q_value"))
    return jsonify(result), result["http_status"]


@app.route("/api/evidence", methods=["GET"])
def evidence_list():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    rows = evidence.list_for_session(sid, kind=request.args.get("kind"))
    return jsonify({"status": "ok", "evidence": rows})


@app.route("/api/evidence/<eid>", methods=["GET"])
def evidence_get(eid):
    row = evidence.get(eid)
    if row is None:
        return jsonify({"error": "unknown evidence"}), 404
    return jsonify({"status": "ok", "evidence": row})


@app.route("/api/evidence/bulk", methods=["POST"])
def evidence_bulk():
    body = request.get_json(force=True)
    rows = evidence.get_bulk(body.get("ids", []))
    return jsonify({"status": "ok", "evidence": rows})


@app.route("/api/evidence/<eid>/review", methods=["POST"])
def evidence_review(eid):
    body = request.get_json(force=True)
    decision = body.get("decision")
    if decision not in ("approved", "rejected"):
        return jsonify({"error": "decision must be 'approved' or 'rejected'"}), 400
    if evidence.get(eid) is None:
        return jsonify({"error": "unknown evidence"}), 404
    if decision == "approved":
        evidence.approve(eid)
    else:
        evidence.reject(eid)
    return jsonify({"status": "ok", "evidence": evidence.get(eid)})


@app.route("/api/evidence/annotate", methods=["POST"])
def evidence_annotate():
    """Human-authored evidence — e.g. a reviewer's own observation, not
    computed by an agent. kind="human_annotation" throughout."""
    body = request.get_json(force=True)
    sid = body.get("session_id")
    claim = body.get("claim")
    if not store.session_exists(sid) or not claim:
        return jsonify({"error": "session_id and claim required"}), 400
    sess = store.get_session(sid)
    row = evidence.from_human_annotation(sid, sess.get("dataset_version_id"), claim,
                                          values=body.get("values"), limitations=body.get("limitations"))
    return jsonify({"status": "ok", "evidence": row})


@app.route("/api/evidence/benchmark", methods=["POST"])
def evidence_benchmark():
    """Takes one already-computed analysis (dashboard) from this session and
    checks it against outside published sources via the Oracle Agent —
    the Evidence stage's own autonomous "look outside the data" action,
    one analysis at a time so a single decline/failure doesn't block the rest."""
    body = request.get_json(force=True)
    sid = body.get("session_id")
    did = body.get("dashboard_id")
    if not store.session_exists(sid) or not did:
        return jsonify({"error": "session_id and dashboard_id required"}), 400
    sess = store.get_session(sid)
    artifact = next((d for d in sess.get("dashboards", []) if d["dashboard_id"] == did), None)
    if artifact is None:
        return jsonify({"error": "unknown dashboard_id"}), 404
    understanding = sess.get("dataset_understanding") or {}
    result = narrative_engine.resolve_evidence_benchmark(
        sid, sess.get("dataset_version_id"), artifact.get("question", artifact.get("title", "")),
        artifact.get("stats", {}), understanding,
    )
    if result.get("status") == "ok" and result.get("evidence_id"):
        artifact["evidence_id"] = result["evidence_id"]
    return jsonify(result)


@app.route("/api/norms", methods=["GET"])
def norms_list():
    return jsonify({"status": "ok", "norms": norm_registry.list_all_current(), "oracle_types": oracle_engine.ORACLE_TYPES})


@app.route("/api/norms/resolve", methods=["POST"])
def norms_resolve():
    body = request.get_json(force=True)
    oracle_type = body.get("oracle_type")
    metric = body.get("metric")
    population = body.get("population") or {}
    if not oracle_type or not metric:
        return jsonify({"error": "oracle_type and metric required"}), 400
    result = norm_registry.resolve_and_register(oracle_type, metric, population)
    return jsonify(result)


@app.route("/api/norms/compare", methods=["POST"])
def norms_compare():
    body = request.get_json(force=True)
    metric = body.get("metric")
    population = body.get("population") or {}
    value = body.get("value")
    sample_size = body.get("sample_size", 0)
    if not metric or value is None:
        return jsonify({"error": "metric and value required"}), 400
    result = norm_registry.compare_to_norm(float(value), metric, population, int(sample_size))
    return jsonify(result)


@app.route("/api/norms/history", methods=["POST"])
def norms_history():
    body = request.get_json(force=True)
    metric = body.get("metric")
    population = body.get("population") or {}
    if not metric:
        return jsonify({"error": "metric required"}), 400
    return jsonify({"status": "ok", "versions": norm_registry.get_version_history(metric, population)})


@app.route("/api/norms/<nid>/approve", methods=["POST"])
def norms_approve(nid):
    if db.get_norm(nid) is None:
        return jsonify({"error": "unknown norm"}), 404
    norm_registry.approve(nid)
    return jsonify({"status": "ok", "norm": db.get_norm(nid)})


@app.route("/api/workspaces", methods=["POST"])
def workspaces_create():
    body = request.get_json(force=True)
    sid = body.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    sess = store.get_session(sid)
    ws = narrative_workspace.create(
        sid, body.get("title", "Untitled narrative"), body.get("thesis", ""),
        body.get("audience", "scientist"), body.get("lens", "efficacy"),
        sess.get("dataset_version_id"), evidence_ids=body.get("evidence_ids"), blocks=body.get("blocks"),
    )
    return jsonify({"status": "ok", "workspace": ws})


@app.route("/api/workspaces", methods=["GET"])
def workspaces_list():
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404
    return jsonify({"status": "ok", "workspaces": narrative_workspace.list_for_session(sid)})


@app.route("/api/workspaces/<wid>", methods=["GET"])
def workspaces_get(wid):
    ws = narrative_workspace.get(wid)
    if ws is None:
        return jsonify({"error": "unknown workspace"}), 404
    return jsonify({"status": "ok", "workspace": ws})


@app.route("/api/workspaces/<wid>/blocks", methods=["POST"])
def workspaces_update_blocks(wid):
    if narrative_workspace.get(wid) is None:
        return jsonify({"error": "unknown workspace"}), 404
    body = request.get_json(force=True)
    ws = narrative_workspace.update_blocks(wid, body.get("blocks", []))
    return jsonify({"status": "ok", "workspace": ws})


@app.route("/api/workspaces/<wid>/meta", methods=["POST"])
def workspaces_update_meta(wid):
    if narrative_workspace.get(wid) is None:
        return jsonify({"error": "unknown workspace"}), 404
    body = request.get_json(force=True)
    ws = narrative_workspace.update_meta(wid, **body)
    return jsonify({"status": "ok", "workspace": ws})


@app.route("/api/workspaces/<wid>/evidence", methods=["POST"])
def workspaces_add_evidence(wid):
    if narrative_workspace.get(wid) is None:
        return jsonify({"error": "unknown workspace"}), 404
    body = request.get_json(force=True)
    eid = body.get("evidence_id")
    if not eid or evidence.get(eid) is None:
        return jsonify({"error": "unknown evidence_id"}), 400
    ws = narrative_workspace.add_evidence(wid, eid)
    return jsonify({"status": "ok", "workspace": ws})


@app.route("/api/workspaces/<wid>/evidence/<eid>", methods=["DELETE"])
def workspaces_remove_evidence(wid, eid):
    if narrative_workspace.get(wid) is None:
        return jsonify({"error": "unknown workspace"}), 404
    ws = narrative_workspace.remove_evidence(wid, eid)
    return jsonify({"status": "ok", "workspace": ws})


@app.route("/api/workspaces/<wid>/transition", methods=["POST"])
def workspaces_transition(wid):
    body = request.get_json(force=True)
    result = narrative_workspace.transition(wid, body.get("status"))
    return jsonify(result), (200 if result["status"] == "ok" else 409)


@app.route("/api/workspaces/<wid>/branch", methods=["POST"])
def workspaces_branch(wid):
    if narrative_workspace.get(wid) is None:
        return jsonify({"error": "unknown workspace"}), 404
    body = request.get_json(force=True)
    branch = narrative_workspace.branch(wid, body.get("title", "Branch"),
                                         body.get("audience", "scientist"), body.get("lens", "efficacy"))
    return jsonify({"status": "ok", "workspace": branch})


@app.route("/api/workspaces/<wid>/branches", methods=["GET"])
def workspaces_branches(wid):
    return jsonify({"status": "ok", "branches": narrative_workspace.list_branches(wid)})


@app.route("/api/workspaces/compare", methods=["GET"])
def workspaces_compare():
    a, b = request.args.get("a"), request.args.get("b")
    result = narrative_workspace.compare(a, b)
    if result is None:
        return jsonify({"error": "unknown workspace(s)"}), 404
    return jsonify({"status": "ok", "comparison": result})


@app.route("/api/workspaces/<wid>/comments", methods=["GET"])
def workspaces_list_comments(wid):
    return jsonify({"status": "ok", "comments": narrative_workspace.list_comments(wid, track=request.args.get("track"))})


@app.route("/api/workspaces/<wid>/comments", methods=["POST"])
def workspaces_add_comment(wid):
    if narrative_workspace.get(wid) is None:
        return jsonify({"error": "unknown workspace"}), 404
    body = request.get_json(force=True)
    result = narrative_workspace.add_comment(wid, body.get("track"), body.get("author", "anonymous"),
                                              body.get("comment", ""), body.get("block_index"))
    return jsonify(result), (200 if result["status"] == "ok" else 400)


@app.route("/api/comments/<cid>/resolve", methods=["POST"])
def comments_resolve(cid):
    narrative_workspace.resolve_comment(cid)
    return jsonify({"status": "ok"})


@app.route("/api/reviews", methods=["GET"])
def reviews_list():
    """Cross-workspace review queue — every open analysis/narrative comment
    across every workspace in the session, plus norm revisions still
    pending_review. Real aggregation over narrative_workspace.py's per-
    workspace comment threads and norm_registry.py's approval_status, not a
    separate review system."""
    sid = request.args.get("session_id")
    if not store.session_exists(sid):
        return jsonify({"error": "unknown session"}), 404

    rows = []
    for ws in narrative_workspace.list_for_session(sid):
        for c in narrative_workspace.list_comments(ws["id"]):
            rows.append({
                "kind": "comment", "id": c["id"], "target": ws["title"], "target_id": ws["id"],
                "track": c["track"], "author": c["author"], "note": c["comment"],
                "status": "resolved" if c["status"] == "resolved" else "open", "created_at": c["created_at"],
            })
    rows.sort(key=lambda r: r["created_at"], reverse=True)
    return jsonify({"status": "ok", "reviews": rows, "open_count": sum(1 for r in rows if r["status"] == "open")})


@app.route("/api/editor/lint", methods=["POST"])
def editor_lint_route():
    body = request.get_json(force=True)
    notes = editor_lint.lint_blocks(body.get("blocks", []))
    return jsonify({"status": "ok", "notes": notes})


if __name__ == "__main__":
    os.makedirs(NOTEBOOK_STATE_ROOT, exist_ok=True)
    app.run(host="0.0.0.0", port=5050, debug=False, use_reloader=False)
